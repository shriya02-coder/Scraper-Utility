import os
import pandas as pd
import openpyxl
import re
from openpyxl.styles import PatternFill

def highlight_rows(file_path, sheet_name, column_name, search_values):
    """
    This Python function highlights rows in an Excel file based on specified search values in a specific
    column.
    
    :param file_path: The `file_path` parameter should be the path to the Excel file you want to work
    with. This should be a string that includes the full path to the file, including the file name and
    extension (e.g., "C:/Users/username/Documents/example.xlsx")
    
    :param sheet_name: The `sheet_name` parameter refers to the name of the specific sheet within the
    Excel file where you want to highlight rows based on the search values. This function will iterate
    through the rows in the specified sheet, check the values in the specified column (if provided), and
    highlight the entire row if the value is present.
    
    :param column_name: The `column_name` parameter in the `highlight_rows` function is used to specify
    the name of the column in the Excel sheet where you want to search for the `search_values` to
    highlight the rows. It is a string that represents the column name in the Excel sheet, such as "A".
    
    :param search_values: Search values are the values that you want to search for in the specified
    column of the Excel file. When a cell in the specified column matches any of these search values,
    the entire row containing that cell will be highlighted in yellow in the Excel file
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    col_index = None
    if column_name:
        col_index = openpyxl.utils.cell.column_index_from_string(column_name)

    for row in sheet.iter_rows():
        for cell in row:
            if col_index is None or cell.column == col_index:
                if str(cell.value) in search_values:
                    for cell in row:
                        cell.fill = yellow_fill
                    break

    dir_path, file_name = os.path.split(file_path)
    new_file_name = 'highlighted_' + file_name
    new_file_path = os.path.join(dir_path, new_file_name)
    workbook.save(new_file_path)

def find_all_matches(text, identifier, extraction_type):
    """
    Find all matches in the given text based on the identifier and extraction type.

    Args:
        text (str): The text to search for matches.
        identifier (str): The identifier to look for in the text.
        extraction_type (str): The type of extraction to perform.

    Returns:
        list: A list of matches found in the text.
    """
    pattern = f"{re.escape(identifier)}[:\\s]+(\\S+)" if extraction_type == "2" else f"{re.escape(identifier)}[:\\s](.)"
    return [match.group(1).strip() for match in re.finditer(pattern, text)]

def process_log_data():
    """
    Process the log data based on user input.
    """
    log_data = []
    print("Please paste the log data (Enter 'END' on a new line when done): ")
    while True:
        line = input()
        if line.strip() == 'END':
            break
        log_data.append(line)

    search_type = input("Do you want to search for a value or an identifier? (1 for value, 2 for identifier) ")

    if search_type == "2":
        identifier = input("Please enter the identifier: ")
        extraction_type = input("Do you want to extract characters or words after the identifier? (1 for characters, 2 for words) ")
        for line in log_data:
            matches = find_all_matches(line, identifier, extraction_type)
            if matches:
                print(f"Line '{line}' -> Matches: {matches}")
    else:
        search_values = [str(value) for value in input("Please enter the values to search for, separated by commas: ").split(',')]
        for line in log_data:
            if any(value in line for value in search_values):
                print(f"Matching Line: {line}")

def main():
    data_source = input("Select data source (1 for Excel, 2 for logs): ")

    if data_source == "1":
        file_path = input("Please enter the Excel file path: ")
        workbook = openpyxl.load_workbook(file_path)
        print("Sheets available:")
        for i, sheet in enumerate(workbook.sheetnames, start=1):
            print(f"{i}. {sheet}")

        sheet_num = input("Enter the number of the sheet you want to search: ")
        sheet_name = workbook.sheetnames[int(sheet_num) - 1]
        sheet = workbook[sheet_name]

        print("\nColumn headers in this sheet:")
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for i, col in enumerate(columns, start=1):
            print(f"{i}. {col}")
        col_choice = input("Enter the column number to search in (0 for all columns): ")
        column_name = openpyxl.utils.cell.get_column_letter(int(col_choice)) if col_choice != "0" else None

        search_values = [str(value) for value in input("Please enter the values to search for, separated by commas: ").split(',')]
        highlight_rows(file_path, sheet_name, column_name, search_values)

    elif data_source == "2":
        process_log_data()

if __name__ == "__main__":
    main()
