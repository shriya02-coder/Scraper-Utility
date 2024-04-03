import os
import pandas as pd
import openpyxl
import re
from openpyxl.styles import PatternFill

def highlight_rows(file_path, sheet_name, column_name, search_values):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    col_index = None
    if column_name:
        col_index = openpyxl.utils.cell.column_index_from_string(column_name)

    for row in sheet.iter_rows():
        for cell in row:
            if col_index is None or cell.column == col_index:
                if str(cell.value) in search_values:
                    for cell in row:
                        cell.fill = yellow_fill
                    break

    dir_path, file_name = os.path.split(file_path)
    new_file_name = 'highlighted_' + file_name
    new_file_path = os.path.join(dir_path, new_file_name)
    workbook.save(new_file_path)

def find_all_matches(text, identifier, num_chars, extraction_type):
    if extraction_type == "1":
        pattern = f"{identifier}(.{{0,{num_chars}}})"
    else:
        pattern = f"{identifier}((?:\\w+\\s*){{0,{num_chars}}})"
    return [match.group(1).strip() for match in re.finditer(pattern, text)]

def process_log_data():
    log_data = []
    print("Please paste the log data (Enter 'END' on a new line when done): ")
    while True:
        line = input()
        if line.strip() == 'END':
            break
        log_data.append(line)

    data_frame = pd.DataFrame(log_data, columns=["Data"])
    search_type = input("Do you want to search for a value or an identifier? (1 for value, 2 for identifier) ")

    if search_type == "2":
        identifiers = input("Please enter the identifiers, separated by commas: ").split(',')
        extraction_type = input("Do you want to extract characters or words after the identifier? (1 for characters, 2 for words) ")
        num = int(input("How many characters/words after the identifier do you want to extract? "))

        data_frame['result'] = data_frame['Data'].apply(lambda x: "\n".join([f"{identifier}: {match}" for identifier in identifiers for match in find_all_matches(x, identifier, num, extraction_type)]))
    else:
        search_values = [str(value) for value in input("Please enter the values to search for, separated by commas: ").split(',')]
        data_frame['result'] = data_frame['Data'].apply(lambda x: [value if value in str(x) else "No such result found" for value in search_values])

    print(data_frame)

def main():
    data_source = input("Select data source (1 for Excel, 2 for logs): ")

    if data_source == "1":
        file_path = input("Please enter the Excel file path: ")
        workbook = openpyxl.load_workbook(file_path)
        print("Sheets available:")
        for i, sheet in enumerate(workbook.sheetnames, start=1):
            print(f"{i}. {sheet}")

        sheet_num = input("Enter the number of the sheet you want to search: ")
        sheet_name = workbook.sheetnames[int(sheet_num) - 1]
        sheet = workbook[sheet_name]

        print("\nColumn headers in this sheet:")
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for i, col in enumerate(columns, start=1):
            print(f"{i}. {col}")
        col_choice = input("Enter the column number to search in (0 for all columns): ")
        column_name = openpyxl.utils.cell.get_column_letter(int(col_choice)) if col_choice != "0" else None

        search_values = [str(value) for value in input("Please enter the values to search for, separated by commas: ").split(',')]
        highlight_rows(file_path, sheet_name, column_name, search_values)

    elif data_source == "2":
        process_log_data()

if __name__ == "__main__":
    main()
